"""Inspect Excel file structure"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\sidda\Desktop\Github Repositories\football-elo\archive\Football-Top5-Past-And-Current-Data.xlsx')
ws = wb.active

# Print headers
print("Column Headers:")
headers = []
for i, cell in enumerate(ws[1], 1):
    headers.append(cell.value)
    print(f"{i}. {cell.value}")

print("\n" + "="*80)
print("First 3 data rows:")
print("="*80)

for row_num in range(2, 5):
    row_data = []
    for cell in ws[row_num]:
        row_data.append(cell.value)

    print(f"\nRow {row_num}:")
    for i, (header, value) in enumerate(zip(headers, row_data), 1):
        print(f"  {i}. {header}: {value}")

# Check last row
print("\n" + "="*80)
print(f"Last row (row {ws.max_row}):")
print("="*80)

last_row = list(ws[ws.max_row])
for i, (header, cell) in enumerate(zip(headers, last_row), 1):
    print(f"  {i}. {header}: {cell.value}")
