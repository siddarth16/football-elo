"""
Football ELO Rating System - Data Processing Script
Processes raw match data and calculates ELO ratings with all custom multipliers
"""

import openpyxl
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
import math

# Constants and Parameters
INITIAL_ELO = 1500
PROMOTED_TEAM_ELO = 1400
BASE_K_FACTOR = 20

# K-Factor Caps (Progressive - allows promoted teams to rise faster)
K_CAPS = {
    1400: 75,   # Elo < 1400
    1500: 60,   # Elo 1400-1500
    1600: 50,   # Elo 1500-1600
    1700: 40,   # Elo 1600-1700
    10000: 35   # Elo > 1700
}

# Multipliers
VENUE_MULTIPLIERS = {
    'away_win': 1.35,
    'away_draw': 1.15,
    'home_win': 1.0,
    'home_draw': 0.9
}

GOAL_DIFFERENCE_MULTIPLIERS = {
    'win': {1: 1.0, 2: 1.15, 3: 1.3, 4: 1.5},
    'loss': {1: 1.0, 2: 0.9, 3: 0.8, 4: 0.7}
}

FORM_MULTIPLIERS = {
    5: 1.2,    # 5 wins
    4: 1.15,   # 4 wins
    3: 1.1,    # 3 wins
    0: 1.0,    # Mixed
    -3: 0.85   # 3+ losses
}

DEFENSIVE_MULTIPLIERS = {
    'clean_sheet_win': 1.15,
    'win_concede_1': 1.0,
    'win_concede_2plus': 0.95,
    'shutout_loss': 0.9
}


class ELOCalculator:
    def __init__(self):
        self.team_elos: Dict[str, float] = {}
        self.team_history: Dict[str, List[Dict]] = defaultdict(list)
        self.match_results: List[Dict] = []

    def get_k_factor_cap(self, elo: float) -> float:
        """Get K-factor cap based on current ELO"""
        for threshold, cap in sorted(K_CAPS.items()):
            if elo < threshold:
                return cap
        return 35

    def calculate_form_score(self, team: str, current_match_idx: int) -> int:
        """Calculate form score based on last 5 games"""
        if not self.team_history[team]:
            return 0

        last_5 = self.team_history[team][-5:]
        wins = sum(1 for m in last_5 if m['result'] == 'W')
        losses = sum(1 for m in last_5 if m['result'] == 'L')

        if wins == 5:
            return 5
        elif wins >= 4:
            return 4
        elif wins >= 3:
            return 3
        elif losses >= 3:
            return -3
        return 0

    def calculate_opponent_quality_multiplier(self, winner_elo: float, loser_elo: float,
                                              is_underdog_win: bool) -> float:
        """Calculate multiplier based on opponent quality"""
        elo_diff = abs(winner_elo - loser_elo)

        if is_underdog_win:
            # Underdog win (lower ELO beats higher ELO)
            # Max multiplier of 2.0 for 400+ ELO difference
            return min(1.0 + (elo_diff / 400), 2.0)
        else:
            # Favorite win
            # Min multiplier of 0.6 for large ELO differences
            return max(1.0 - (elo_diff / 800), 0.6)

    def calculate_venue_multiplier(self, is_home: bool, result: str) -> float:
        """Calculate venue-based multiplier"""
        if is_home:
            return VENUE_MULTIPLIERS['home_win'] if result == 'W' else VENUE_MULTIPLIERS['home_draw']
        else:
            return VENUE_MULTIPLIERS['away_win'] if result == 'W' else VENUE_MULTIPLIERS['away_draw']

    def calculate_gd_multiplier(self, goal_diff: int, is_winner: bool) -> float:
        """Calculate goal difference multiplier"""
        abs_gd = abs(goal_diff)

        if is_winner:
            # Cap at 4+ goals
            gd_key = min(abs_gd, 4)
            return GOAL_DIFFERENCE_MULTIPLIERS['win'].get(gd_key, 1.5)
        else:
            # Losses
            gd_key = min(abs_gd, 4)
            return GOAL_DIFFERENCE_MULTIPLIERS['loss'].get(gd_key, 0.7)

    def calculate_form_multiplier(self, team: str, current_match_idx: int) -> float:
        """Calculate form-based multiplier"""
        form_score = self.calculate_form_score(team, current_match_idx)
        return FORM_MULTIPLIERS.get(form_score, 1.0)

    def calculate_defensive_multiplier(self, goals_scored: int, goals_conceded: int,
                                      result: str) -> float:
        """Calculate defensive quality multiplier"""
        if result == 'W':
            if goals_conceded == 0:
                return DEFENSIVE_MULTIPLIERS['clean_sheet_win']
            elif goals_conceded == 1:
                return DEFENSIVE_MULTIPLIERS['win_concede_1']
            else:
                return DEFENSIVE_MULTIPLIERS['win_concede_2plus']
        elif result == 'L' and goals_scored == 0:
            return DEFENSIVE_MULTIPLIERS['shutout_loss']
        return 1.0

    def calculate_expected_score(self, home_elo: float, away_elo: float,
                                home_advantage: float = 50) -> Tuple[float, float]:
        """Calculate expected win probability using ELO formula"""
        expected_home = 1 / (1 + 10 ** ((away_elo - home_elo - home_advantage) / 400))
        expected_away = 1 - expected_home
        return expected_home, expected_away

    def calculate_elo_change(self, team: str, opponent: str, is_home: bool,
                            goals_scored: int, goals_conceded: int,
                            result: str, current_match_idx: int,
                            home_advantage: float = 50) -> Tuple[float, Dict]:
        """
        Calculate ELO change for a team based on match result
        Returns: (elo_change, multipliers_dict)
        """
        team_elo = self.team_elos.get(team, INITIAL_ELO)
        opponent_elo = self.team_elos.get(opponent, INITIAL_ELO)

        # Calculate expected score
        if is_home:
            expected, _ = self.calculate_expected_score(team_elo, opponent_elo, home_advantage)
        else:
            _, expected = self.calculate_expected_score(opponent_elo, team_elo, home_advantage)

        # Actual score
        if result == 'W':
            actual = 1.0
            # GD enhancement for wins
            gd = abs(goals_scored - goals_conceded)
            if gd >= 4:
                actual = 1.3
            elif gd == 3:
                actual = 1.2
            elif gd == 2:
                actual = 1.1
        elif result == 'D':
            actual = 0.5
        else:
            actual = 0.0

        # Calculate all multipliers
        multipliers = {}

        # 1. Opponent Quality
        if result == 'W':
            is_underdog = team_elo < opponent_elo
            multipliers['opponent'] = self.calculate_opponent_quality_multiplier(
                team_elo, opponent_elo, is_underdog
            )
        else:
            multipliers['opponent'] = 1.0

        # 2. Venue
        multipliers['venue'] = self.calculate_venue_multiplier(is_home, result)

        # 3. Goal Difference
        gd = goals_scored - goals_conceded
        multipliers['gd'] = self.calculate_gd_multiplier(gd, result == 'W')

        # 4. Form
        multipliers['form'] = self.calculate_form_multiplier(team, current_match_idx)

        # 5. Defense
        multipliers['defense'] = self.calculate_defensive_multiplier(
            goals_scored, goals_conceded, result
        )

        # Calculate base K
        base_k = BASE_K_FACTOR

        # Apply all multipliers
        k_adjusted = base_k
        for mult_name, mult_value in multipliers.items():
            k_adjusted *= mult_value

        # Apply K-factor cap
        k_cap = self.get_k_factor_cap(team_elo)
        k_final = min(k_adjusted, k_cap)

        # Calculate ELO change
        elo_change = k_final * (actual - expected)

        return elo_change, {
            'k_base': base_k,
            'k_adjusted': k_adjusted,
            'k_final': k_final,
            'k_cap': k_cap,
            'expected': expected,
            'actual': actual,
            **multipliers
        }

    def process_match(self, match_data: Dict, match_idx: int, home_advantage: float = 50) -> Dict:
        """Process a single match and update ELOs"""
        home_team = match_data['homeTeamName']
        away_team = match_data['awayTeamName']
        home_score = match_data['homeTeamScore']
        away_score = match_data['awayTeamScore']

        # Skip if scores are None (future matches)
        if home_score is None or away_score is None:
            return None

        # Initialize ELOs if not present
        if home_team not in self.team_elos:
            self.team_elos[home_team] = INITIAL_ELO
        if away_team not in self.team_elos:
            self.team_elos[away_team] = INITIAL_ELO

        # Get pre-match ELOs
        home_elo_pre = self.team_elos[home_team]
        away_elo_pre = self.team_elos[away_team]

        # Determine results
        if home_score > away_score:
            home_result = 'W'
            away_result = 'L'
        elif home_score < away_score:
            home_result = 'L'
            away_result = 'W'
        else:
            home_result = 'D'
            away_result = 'D'

        # Calculate ELO changes
        home_elo_change, home_mult = self.calculate_elo_change(
            home_team, away_team, True, home_score, away_score, home_result, match_idx, home_advantage
        )
        away_elo_change, away_mult = self.calculate_elo_change(
            away_team, home_team, False, away_score, home_score, away_result, match_idx, home_advantage
        )

        # Update ELOs
        self.team_elos[home_team] += home_elo_change
        self.team_elos[away_team] += away_elo_change

        # Record in history
        self.team_history[home_team].append({
            'result': home_result,
            'opponent': away_team,
            'elo_change': home_elo_change
        })
        self.team_history[away_team].append({
            'result': away_result,
            'opponent': home_team,
            'elo_change': away_elo_change
        })

        # Create match result record
        match_result = {
            **match_data,
            'home_elo_pre': home_elo_pre,
            'away_elo_pre': away_elo_pre,
            'home_elo_post': self.team_elos[home_team],
            'away_elo_post': self.team_elos[away_team],
            'home_elo_change': home_elo_change,
            'away_elo_change': away_elo_change,
            'home_result': home_result,
            'away_result': away_result,
            'goal_diff': home_score - away_score,
            'home_multipliers': home_mult,
            'away_multipliers': away_mult
        }

        self.match_results.append(match_result)
        return match_result


def load_raw_data(file_path: str) -> List[Dict]:
    """Load raw data from Excel file"""
    print(f"Loading data from {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Super Data']

    matches = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        match_data = dict(zip(headers, row))
        matches.append(match_data)

    wb.close()
    print(f"Loaded {len(matches)} matches")
    return matches


def calculate_baseline_stats(matches_2024: List[Dict]) -> Dict:
    """Calculate baseline statistics from 2024-25 season"""
    print("\nCalculating baseline statistics from 2024-25 season...")

    total_matches = len([m for m in matches_2024 if m['homeTeamScore'] is not None])
    draws = len([m for m in matches_2024 if m['homeTeamScore'] == m['awayTeamScore']
                 and m['homeTeamScore'] is not None])
    home_wins = len([m for m in matches_2024 if m['homeTeamWinner'] == True])
    away_wins = len([m for m in matches_2024 if m['awayTeamWinner'] == True])

    # Team-specific home advantage stats
    team_home_stats = defaultdict(lambda: {'wins': 0, 'matches': 0})
    team_away_stats = defaultdict(lambda: {'wins': 0, 'matches': 0})
    team_defensive_stats = defaultdict(lambda: {'clean_sheets': 0, 'goals_conceded': 0, 'matches': 0})

    for match in matches_2024:
        if match['homeTeamScore'] is None:
            continue

        home_team = match['homeTeamName']
        away_team = match['awayTeamName']

        # Home stats
        team_home_stats[home_team]['matches'] += 1
        if match['homeTeamWinner']:
            team_home_stats[home_team]['wins'] += 1

        # Away stats
        team_away_stats[away_team]['matches'] += 1
        if match['awayTeamWinner']:
            team_away_stats[away_team]['wins'] += 1

        # Defensive stats
        team_defensive_stats[home_team]['matches'] += 1
        team_defensive_stats[home_team]['goals_conceded'] += match['awayTeamScore']
        if match['awayTeamScore'] == 0:
            team_defensive_stats[home_team]['clean_sheets'] += 1

        team_defensive_stats[away_team]['matches'] += 1
        team_defensive_stats[away_team]['goals_conceded'] += match['homeTeamScore']
        if match['homeTeamScore'] == 0:
            team_defensive_stats[away_team]['clean_sheets'] += 1

    # Calculate averages
    draw_percentage = (draws / total_matches) * 100
    home_win_percentage = (home_wins / total_matches) * 100
    away_win_percentage = (away_wins / total_matches) * 100

    # Calculate average home advantage
    home_advantages = {}
    for team, stats in team_home_stats.items():
        if stats['matches'] > 0:
            home_win_rate = stats['wins'] / stats['matches']
            away_stats = team_away_stats[team]
            away_win_rate = away_stats['wins'] / away_stats['matches'] if away_stats['matches'] > 0 else 0
            # Convert to ELO points (rough estimation)
            home_advantages[team] = 30 + (home_win_rate * 40)

    avg_home_advantage = sum(home_advantages.values()) / len(home_advantages) if home_advantages else 50

    # Defensive quality
    defensive_quality = {}
    for team, stats in team_defensive_stats.items():
        if stats['matches'] > 0:
            clean_sheet_rate = stats['clean_sheets'] / stats['matches']
            avg_goals_conceded = stats['goals_conceded'] / stats['matches']
            defensive_quality[team] = {
                'clean_sheet_rate': clean_sheet_rate,
                'avg_goals_conceded': avg_goals_conceded,
                'defensive_score': clean_sheet_rate * 0.7 + (1 - min(avg_goals_conceded / 2, 1)) * 0.3
            }

    baseline = {
        'draw_percentage': draw_percentage,
        'home_win_percentage': home_win_percentage,
        'away_win_percentage': away_win_percentage,
        'avg_home_advantage': avg_home_advantage,
        'team_home_advantages': home_advantages,
        'team_defensive_quality': defensive_quality
    }

    print(f"  Draw %: {draw_percentage:.2f}%")
    print(f"  Home Win %: {home_win_percentage:.2f}%")
    print(f"  Away Win %: {away_win_percentage:.2f}%")
    print(f"  Avg Home Advantage: {avg_home_advantage:.1f} ELO points")

    return baseline


def main():
    """Main processing function"""
    print("="*80)
    print("FOOTBALL ELO RATING SYSTEM - DATA PROCESSING")
    print("="*80)

    # Load data
    raw_file = r"C:\Users\sidda\Desktop\Github Repositories\football-elo\Football-Top5-Past-And-Current-Data.xlsx"
    all_matches = load_raw_data(raw_file)

    # Split by season
    matches_2024 = [m for m in all_matches if '2024-25' in str(m['seasonName'])]
    matches_2025 = [m for m in all_matches if '2025-26' in str(m['seasonName'])]

    print(f"\n2024-25 Season: {len(matches_2024)} matches")
    print(f"2025-26 Season: {len(matches_2025)} matches")

    # Calculate baseline stats
    baseline_stats = calculate_baseline_stats(matches_2024)

    # Process 2024-25 season
    print("\n" + "="*80)
    print("PROCESSING 2024-25 SEASON (Training Data)")
    print("="*80)

    calculator = ELOCalculator()

    # Sort by date
    matches_2024_sorted = sorted(matches_2024, key=lambda x: x['date'])

    processed_2024 = []
    for idx, match in enumerate(matches_2024_sorted):
        result = calculator.process_match(match, idx, baseline_stats['avg_home_advantage'])
        if result:
            processed_2024.append(result)

        if (idx + 1) % 100 == 0:
            print(f"  Processed {idx + 1}/{len(matches_2024_sorted)} matches...")

    print(f"\nCompleted! Processed {len(processed_2024)} matches from 2024-25 season")

    # Get final ELOs for 2024-25
    final_elos_2024 = calculator.team_elos.copy()

    print("\nTop 10 Teams (End of 2024-25):")
    sorted_teams = sorted(final_elos_2024.items(), key=lambda x: x[1], reverse=True)[:10]
    for rank, (team, elo) in enumerate(sorted_teams, 1):
        print(f"  {rank:2d}. {team:30s}: {elo:.1f}")

    # Save processed 2024-25 data
    output_2024 = {
        'matches': processed_2024,
        'final_elos': final_elos_2024,
        'baseline_stats': baseline_stats
    }

    output_file_2024 = r"C:\Users\sidda\Desktop\Github Repositories\football-elo\football-elo-webapp\data\season_2024_25.json"
    with open(output_file_2024, 'w', encoding='utf-8') as f:
        json.dump(output_2024, f, indent=2, default=str)
    print(f"\nSaved 2024-25 season data to {output_file_2024}")

    # Prepare 2025-26 season data (matches with and without scores)
    print("\n" + "="*80)
    print("PREPARING 2025-26 SEASON DATA")
    print("="*80)

    matches_2025_sorted = sorted(matches_2025, key=lambda x: x['date'])

    # Create new calculator with final 2024-25 ELOs as starting point
    calculator_2025 = ELOCalculator()
    calculator_2025.team_elos = final_elos_2024.copy()

    # Identify promoted teams (teams in 2025-26 but not in 2024-25)
    teams_2024 = set(final_elos_2024.keys())
    teams_2025 = set([m['homeTeamName'] for m in matches_2025] + [m['awayTeamName'] for m in matches_2025])
    promoted_teams = teams_2025 - teams_2024

    # Set promoted teams to promoted ELO
    for team in promoted_teams:
        calculator_2025.team_elos[team] = PROMOTED_TEAM_ELO

    print(f"\nPromoted teams ({len(promoted_teams)}):")
    for team in sorted(promoted_teams):
        print(f"  - {team} (Starting ELO: {PROMOTED_TEAM_ELO})")

    # Process matches that have scores
    processed_2025 = []
    pending_2025 = []

    for idx, match in enumerate(matches_2025_sorted):
        if match['homeTeamScore'] is not None and match['awayTeamScore'] is not None:
            result = calculator_2025.process_match(match, idx, baseline_stats['avg_home_advantage'])
            if result:
                processed_2025.append(result)
        else:
            # Store as pending match
            pending_match = {
                **match,
                'home_elo_current': calculator_2025.team_elos.get(match['homeTeamName'], INITIAL_ELO),
                'away_elo_current': calculator_2025.team_elos.get(match['awayTeamName'], INITIAL_ELO)
            }
            pending_2025.append(pending_match)

    print(f"\nProcessed {len(processed_2025)} completed matches")
    print(f"Pending {len(pending_2025)} upcoming matches")

    # Save 2025-26 data
    output_2025 = {
        'completed_matches': processed_2025,
        'pending_matches': pending_2025,
        'current_elos': calculator_2025.team_elos,
        'promoted_teams': list(promoted_teams)
    }

    output_file_2025 = r"C:\Users\sidda\Desktop\Github Repositories\football-elo\football-elo-webapp\data\season_2025_26.json"
    with open(output_file_2025, 'w', encoding='utf-8') as f:
        json.dump(output_2025, f, indent=2, default=str)
    print(f"Saved 2025-26 season data to {output_file_2025}")

    # Save parameters
    params = {
        'initial_elo': INITIAL_ELO,
        'promoted_team_elo': PROMOTED_TEAM_ELO,
        'base_k_factor': BASE_K_FACTOR,
        'k_caps': K_CAPS,
        'venue_multipliers': VENUE_MULTIPLIERS,
        'gd_multipliers': GOAL_DIFFERENCE_MULTIPLIERS,
        'form_multipliers': FORM_MULTIPLIERS,
        'defensive_multipliers': DEFENSIVE_MULTIPLIERS,
        'baseline_stats': baseline_stats
    }

    params_file = r"C:\Users\sidda\Desktop\Github Repositories\football-elo\football-elo-webapp\data\parameters.json"
    with open(params_file, 'w', encoding='utf-8') as f:
        json.dump(params, f, indent=2, default=str)
    print(f"Saved parameters to {params_file}")

    print("\n" + "="*80)
    print("DATA PROCESSING COMPLETE!")
    print("="*80)


if __name__ == "__main__":
    main()
