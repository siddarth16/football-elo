"""Check what data exists in the Excel file"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\sidda\Desktop\Github Repositories\football-elo\archive\Football-Top5-Past-And-Current-Data.xlsx')
print(f"Sheets: {wb.sheetnames}")

ws = wb.active
print(f"\nTotal rows: {ws.max_row}")

# Check date range
dates = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
    if row[0].value:
        dates.append(row[0].value)

if dates:
    print(f"\nFirst match date: {dates[0]}")
    print(f"Last match date: {dates[-1]}")
    print(f"Total matches in Excel: {len(dates)}")

# Count 2025-26 season matches
season_2025_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[3] == 2025:  # seasonYear column
        season_2025_count += 1

print(f"\n2025-26 season matches in Excel: {season_2025_count}")

# Check for matches after Oct 5, 2025
future_matches = []
cutoff = datetime(2025, 10, 5, 23, 59, 59)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    match_date = row[7]  # date column
    if isinstance(match_date, datetime) and match_date > cutoff and row[3] == 2025:
        future_matches.append(row)

print(f"\nMatches after Oct 5, 2025 in Excel: {len(future_matches)}")

if future_matches:
    print("\nFirst 5 future matches:")
    for i, match in enumerate(future_matches[:5], 1):
        print(f"{i}. {match[7]} - {match[12]} vs {match[14]}")
