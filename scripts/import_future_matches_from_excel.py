"""
Import all future matches from Excel file starting from Oct 17, 2025
"""

import os
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
from supabase import create_client

load_dotenv('.env.local')

supabase = create_client(
    os.getenv('NEXT_PUBLIC_SUPABASE_URL'),
    os.getenv('SUPABASE_SERVICE_KEY')
)

print("="*80)
print("IMPORTING FUTURE MATCHES FROM EXCEL")
print("="*80)

# Load Excel file
wb = openpyxl.load_workbook(
    r'C:\Users\sidda\Desktop\Github Repositories\football-elo\archive\Football-Top5-Past-And-Current-Data.xlsx',
    data_only=True  # Read calculated values, not formulas
)
ws = wb.active

print(f"\nTotal rows in Excel: {ws.max_row}")

# Get column indices
headers = [cell.value for cell in ws[1]]
print(f"Columns: {len(headers)}")

# Cutoff date - Oct 17, 2025
cutoff_date = datetime(2025, 10, 17, 0, 0, 0)

# First, delete the sample matches I added earlier
print("\nDeleting sample matches...")
delete_result = supabase.table('matches').delete().gte('event_id', 800000).execute()
print(f"Deleted sample matches")

# Also delete their predictions
supabase.table('predictions').delete().gte('event_id', 800000).execute()

# Get existing event IDs in database to avoid duplicates
existing_matches = supabase.table('matches').select('event_id').execute()
existing_event_ids = {m['event_id'] for m in existing_matches.data}

print(f"\nExisting matches in database: {len(existing_event_ids)}")

# Read all matches from Excel starting Oct 17, 2025
future_matches = []
skipped = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    event_id = row[6]  # eventId
    match_date = row[7]  # date

    # Skip if not a datetime or before cutoff
    if not isinstance(match_date, datetime) or match_date < cutoff_date:
        continue

    # Skip if already in database
    if event_id in existing_event_ids:
        skipped += 1
        continue

    # Skip if it has a score (0-0 with statusId 1 means not played)
    home_score = row[16]  # homeTeamScore
    away_score = row[17]  # awayTeamScore
    status_id = row[20]  # statusId

    # Only include if statusId indicates not played (1 = scheduled)
    # Skip if it has actual scores entered (not 0-0 placeholder)
    if status_id != 1:
        continue

    future_matches.append({
        'event_id': event_id,
        'season_type': row[1],  # seasonType
        'season_name': row[2],  # seasonName
        'season_year': row[3],  # seasonYear
        'league_id': row[4],  # leagueId
        'league_name': row[5],  # leagueName
        'match_date': match_date.isoformat(),
        'venue_id': row[8],  # venueId
        'attendance': row[9],  # attendance
        'home_team_id': row[10],  # homeTeamId
        'home_team_name': row[11],  # homeTeamName
        'away_team_id': row[12],  # awayTeamId
        'away_team_name': row[13],  # awayTeamName
        'is_completed': False,
        'home_team_score': None,
        'away_team_score': None,
        'home_team_winner': None,
        'away_team_winner': None
    })

print(f"\nFound {len(future_matches)} new future matches to import")
print(f"Skipped {skipped} matches already in database")

if future_matches:
    print(f"\nFirst 10 matches to import:")
    for i, match in enumerate(future_matches[:10], 1):
        print(f"{i}. {match['match_date'][:10]}: {match['home_team_name']} vs {match['away_team_name']} ({match['league_name']})")

    # Insert matches in batches of 100
    batch_size = 100
    total_inserted = 0

    for i in range(0, len(future_matches), batch_size):
        batch = future_matches[i:i+batch_size]
        try:
            supabase.table('matches').insert(batch).execute()
            total_inserted += len(batch)
            print(f"\nInserted batch {i//batch_size + 1}: {len(batch)} matches (Total: {total_inserted})")
        except Exception as e:
            print(f"\nError inserting batch {i//batch_size + 1}: {e}")
            # Try inserting one by one
            for match in batch:
                try:
                    supabase.table('matches').insert([match]).execute()
                    total_inserted += 1
                except Exception as e2:
                    print(f"  Failed to insert match {match['event_id']}: {e2}")

    print(f"\nTotal matches inserted: {total_inserted}")

    # Now regenerate ALL predictions
    print("\n" + "="*80)
    print("REGENERATING ALL PREDICTIONS")
    print("="*80)

    # Get home advantage
    params = supabase.table('parameters').select('*').execute()
    params_dict = {p['param_key']: p['param_value'] for p in params.data}
    home_advantage = params_dict.get('baseline_stats', {}).get('avg_home_advantage', 46.8)

    # Get current ELOs
    teams = supabase.table('teams').select('name, current_elo').execute()
    current_elos = {team['name']: team['current_elo'] for team in teams.data}

    # Get all pending matches
    pending_matches = supabase.table('matches') \
        .select('*') \
        .eq('is_completed', False) \
        .order('match_date', desc=False) \
        .execute()

    print(f"\nTotal pending matches: {len(pending_matches.data)}")

    # Delete all existing predictions
    supabase.table('predictions').delete().neq('id', 0).execute()

    # Generate predictions
    predictions = []
    for match in pending_matches.data:
        home_team = match['home_team_name']
        away_team = match['away_team_name']
        home_elo = current_elos.get(home_team, 1500)
        away_elo = current_elos.get(away_team, 1500)

        # Calculate prediction
        expected_home = 1 / (1 + 10 ** ((away_elo - home_elo - home_advantage) / 400))
        base_draw = 0.2494
        elo_diff = abs(home_elo - away_elo)
        closeness_bonus = max(0, (200 - min(elo_diff, 200)) / 2000)
        draw_prob = max(0.15, min(0.40, base_draw * (1 + closeness_bonus)))

        remaining = 1 - draw_prob
        home_win_prob = expected_home * remaining
        away_win_prob = (1 - expected_home) * remaining

        home_or_draw = home_win_prob + draw_prob
        away_or_draw = away_win_prob + draw_prob

        # Determine recommendation
        if home_win_prob >= 0.40:
            recommended = ('Home Win', home_win_prob)
        elif away_win_prob >= 0.40:
            recommended = ('Away Win', away_win_prob)
        elif draw_prob >= 0.40:
            recommended = ('Draw', draw_prob)
        elif home_or_draw > 0.60:
            recommended = ('Home Win/Draw', home_or_draw)
        elif away_or_draw > 0.60:
            recommended = ('Away Win/Draw', away_or_draw)
        else:
            recommended = ('Home Win' if home_win_prob > away_win_prob else 'Away Win',
                          max(home_win_prob, away_win_prob))

        predictions.append({
            'event_id': match['event_id'],
            'match_id': match['id'],
            'home_elo': home_elo,
            'away_elo': away_elo,
            'home_win_prob': round(home_win_prob, 4),
            'draw_prob': round(draw_prob, 4),
            'away_win_prob': round(away_win_prob, 4),
            'home_or_draw_prob': round(home_or_draw, 4),
            'away_or_draw_prob': round(away_or_draw, 4),
            'recommended_bet': recommended[0],
            'recommended_prob': round(recommended[1], 4),
            'confidence': 'High' if recommended[1] > 0.6 else ('Medium' if recommended[1] > 0.5 else 'Low')
        })

    # Insert predictions in batches
    if predictions:
        for i in range(0, len(predictions), batch_size):
            batch = predictions[i:i+batch_size]
            supabase.table('predictions').insert(batch).execute()
            print(f"Inserted prediction batch {i//batch_size + 1}: {len(batch)} predictions")

        print(f"\nTotal predictions generated: {len(predictions)}")

        # Show sample
        print("\nSample predictions (first 5):")
        for i, pred in enumerate(predictions[:5], 1):
            match = next(m for m in pending_matches.data if m['event_id'] == pred['event_id'])
            print(f"\n{i}. {match['match_date'][:10]}: {match['home_team_name']} vs {match['away_team_name']}")
            print(f"   Home: {pred['home_win_prob']*100:.1f}% | Draw: {pred['draw_prob']*100:.1f}% | Away: {pred['away_win_prob']*100:.1f}%")
            print(f"   Recommended: {pred['recommended_bet']} ({pred['recommended_prob']*100:.1f}%) - {pred['confidence']}")

else:
    print("\nNo new matches to import")

print("\n" + "="*80)
print("IMPORT COMPLETE")
print("="*80)
